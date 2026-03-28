#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Fonctions d'export pour ZKTeco Manager
Export Excel et PDF
"""

import os
from datetime import datetime, date, timedelta
from typing import List, Dict, Optional

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


def export_attendance_to_excel(attendance: List[Dict], filename: str) -> bool:
    """
    Exporter les pointages vers un fichier Excel.
    
    Args:
        attendance: Liste des pointages
        filename: Chemin du fichier de sortie
    
    Returns:
        bool: True si succès
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl n'est pas installé")
    
    # Créer le workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Pointages"
    
    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-têtes
    headers = ['ID', 'Employé', 'Date', 'Heure', 'Status', 'Vérification', 'Terminal']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Données
    status_names = {
        0: "Entrée", 1: "Sortie", 2: "Sortie Pause", 3: "Retour Pause",
        4: "Entrée H.Sup", 5: "Sortie H.Sup"
    }
    
    verify_names = {
        0: "Mot de passe", 1: "Empreinte", 2: "Carte", 3: "Visage"
    }
    
    for row_idx, record in enumerate(attendance, 2):
        ts = record.get('timestamp')
        if isinstance(ts, datetime):
            date_str = ts.strftime('%d/%m/%Y')
            time_str = ts.strftime('%H:%M:%S')
        else:
            date_str = str(ts)
            time_str = ''
        
        ws.cell(row=row_idx, column=1, value=record.get('id'))
        ws.cell(row=row_idx, column=2, value=record.get('user_name', ''))
        ws.cell(row=row_idx, column=3, value=date_str)
        ws.cell(row=row_idx, column=4, value=time_str)
        ws.cell(row=row_idx, column=5, value=status_names.get(record.get('status', 0), 'N/A'))
        ws.cell(row=row_idx, column=6, value=verify_names.get(record.get('verify_type', 0), 'N/A'))
        ws.cell(row=row_idx, column=7, value=record.get('terminal_id', ''))
        
        # Appliquer les bordures
        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).border = border
    
    # Ajuster les largeurs
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    # Sauvegarder
    wb.save(filename)
    
    return True


def export_users_to_excel(users: List[Dict], filename: str) -> bool:
    """
    Exporter les utilisateurs vers un fichier Excel.
    
    Args:
        users: Liste des utilisateurs
        filename: Chemin du fichier de sortie
    
    Returns:
        bool: True si succès
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl n'est pas installé")
    
    # Créer le workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Utilisateurs"
    
    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-têtes
    headers = ['ID', 'UID', 'Nom', 'Privilège', 'Département', 'Carte', 'Actif']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Données
    privilege_names = {0: "Utilisateur", 2: "Enregistreur", 3: "Gestionnaire", 6: "Super Admin"}
    
    for row_idx, user in enumerate(users, 2):
        ws.cell(row=row_idx, column=1, value=user.get('id'))
        ws.cell(row=row_idx, column=2, value=user.get('uid'))
        ws.cell(row=row_idx, column=3, value=user.get('name'))
        ws.cell(row=row_idx, column=4, value=privilege_names.get(user.get('privilege', 0), 'Utilisateur'))
        ws.cell(row=row_idx, column=5, value=user.get('department_name', ''))
        ws.cell(row=row_idx, column=6, value=user.get('card', '') or '')
        ws.cell(row=row_idx, column=7, value="Oui" if user.get('is_active', 1) else "Non")
        
        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).border = border
    
    # Ajuster les largeurs
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 8
    
    wb.save(filename)
    
    return True


def export_attendance_to_pdf(attendance: List[Dict], filename: str, 
                            title: str = "Rapport de pointages",
                            company_name: str = "") -> bool:
    """
    Exporter les pointages vers un fichier PDF.
    
    Args:
        attendance: Liste des pointages
        filename: Chemin du fichier de sortie
        title: Titre du rapport
        company_name: Nom de l'entreprise
    
    Returns:
        bool: True si succès
    """
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab n'est pas installé")
    
    # Créer le document
    doc = SimpleDocTemplate(
        filename,
        pagesize=landscape(A4),
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    # Styles
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=20,
        alignment=1  # Centré
    )
    
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=10,
        alignment=1
    )
    
    # Contenu
    elements = []
    
    # Titre
    if company_name:
        elements.append(Paragraph(company_name, subtitle_style))
    
    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(
        f"Période: {datetime.now().strftime('%d/%m/%Y')}",
        subtitle_style
    ))
    elements.append(Spacer(1, 20))
    
    # Tableau
    headers = ['ID', 'Employé', 'Date', 'Heure', 'Status', 'Vérification']
    
    status_names = {
        0: "Entrée", 1: "Sortie", 2: "Sortie Pause", 3: "Retour Pause",
        4: "Entrée H.Sup", 5: "Sortie H.Sup"
    }
    
    verify_names = {
        0: "Mot de passe", 1: "Empreinte", 2: "Carte", 3: "Visage"
    }
    
    data = [headers]
    
    for record in attendance[:100]:  # Limiter à 100 pour le PDF
        ts = record.get('timestamp')
        if isinstance(ts, datetime):
            date_str = ts.strftime('%d/%m/%Y')
            time_str = ts.strftime('%H:%M:%S')
        else:
            date_str = str(ts)
            time_str = ''
        
        data.append([
            str(record.get('id', '')),
            str(record.get('user_name', '')),
            date_str,
            time_str,
            status_names.get(record.get('status', 0), 'N/A'),
            verify_names.get(record.get('verify_type', 0), 'N/A')
        ])
    
    table = Table(data, repeatRows=1)
    
    table.setStyle(TableStyle([
        # En-têtes
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        
        # Données
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        
        # Bordures
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        
        # Alternance des couleurs
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
    ]))
    
    elements.append(table)
    
    # Pied de page
    elements.append(Spacer(1, 30))
    elements.append(Paragraph(
        f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M:%S')} - ZKTeco Manager",
        styles['Normal']
    ))
    
    # Générer le PDF
    doc.build(elements)
    
    return True


def generate_report_pdf(report_type: str, data: Dict, filename: str,
                       company_name: str = "") -> bool:
    """
    Générer un rapport PDF.
    
    Args:
        report_type: Type de rapport ('daily', 'monthly', 'summary')
        data: Données du rapport
        filename: Chemin du fichier de sortie
        company_name: Nom de l'entreprise
    
    Returns:
        bool: True si succès
    """
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab n'est pas installé")
    
    # Créer le document
    doc = SimpleDocTemplate(
        filename,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    # Styles
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=15,
        alignment=1
    )
    
    # Contenu
    elements = []
    
    # Titre
    if company_name:
        elements.append(Paragraph(company_name, styles['Heading2']))
    
    if report_type == 'daily':
        elements.append(Paragraph("Rapport de présence journalière", title_style))
    elif report_type == 'monthly':
        elements.append(Paragraph("Rapport de présence mensuelle", title_style))
    else:
        elements.append(Paragraph("Rapport de synthèse", title_style))
    
    elements.append(Spacer(1, 20))
    
    # Résumé
    summary_data = data.get('summary', {})
    
    if summary_data:
        summary_text = f"""
        <b>Résumé:</b><br/>
        - Total employés: {summary_data.get('total_users', 0)}<br/>
        - Présents: {summary_data.get('present', 0)}<br/>
        - Absents: {summary_data.get('absent', 0)}<br/>
        """
        elements.append(Paragraph(summary_text, styles['Normal']))
        elements.append(Spacer(1, 20))
    
    # Détails
    details = data.get('details', [])
    if details:
        # Créer un tableau avec les détails
        headers = ['Employé', 'Première entrée', 'Dernière sortie', 'Heures', 'Status']
        
        table_data = [headers]
        for detail in details:
            table_data.append([
                str(detail.get('name', '')),
                str(detail.get('first_in', '-')),
                str(detail.get('last_out', '-')),
                str(detail.get('hours', '-')),
                str(detail.get('status', ''))
            ])
        
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ]))
        
        elements.append(table)
    
    # Pied de page
    elements.append(Spacer(1, 30))
    elements.append(Paragraph(
        f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M:%S')} - ZKTeco Manager",
        styles['Normal']
    ))
    
    doc.build(elements)
    
    return True


def export_to_csv(data: List[Dict], filename: str, headers: List[str] = None) -> bool:
    """
    Exporter les données vers un fichier CSV.
    
    Args:
        data: Liste des enregistrements
        filename: Chemin du fichier de sortie
        headers: Liste des colonnes à exporter
    
    Returns:
        bool: True si succès
    """
    import csv
    
    if not data:
        return False
    
    # Déterminer les en-têtes
    if headers is None:
        headers = list(data[0].keys())
    
    with open(filename, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(data)
    
    return True


def create_backup(db_path: str, backup_dir: str) -> str:
    """
    Créer une sauvegarde de la base de données.
    
    Args:
        db_path: Chemin de la base de données
        backup_dir: Dossier de sauvegarde
    
    Returns:
        str: Chemin du fichier de sauvegarde
    """
    import shutil
    
    os.makedirs(backup_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_name = f"zkteco_backup_{timestamp}.db"
    backup_path = os.path.join(backup_dir, backup_name)
    
    shutil.copy2(db_path, backup_path)
    
    return backup_path
